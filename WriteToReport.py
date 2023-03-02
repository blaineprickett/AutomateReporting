import openpyxl


def copy_data():
    # open the source workbooks
    workbook1 = openpyxl.load_workbook('C:\\Users\\Cprickett\\')
    workbook2 = openpyxl.load_workbook('C:\\Users\\Cprickett\\')
    workbook3 = openpyxl.load_workbook('C:\\Users\\Cprickett\\')

    workbook4 = openpyxl.load_workbook('C:\\Users\\Cprickett\\')
    workbook5 = openpyxl.load_workbook('C:\\Users\\Cprickett\\')
    workbook6 = openpyxl.load_workbook('C:\\Users\\Cprickett\\')

    # get the source worksheets
    worksheet1 = workbook1['Sheet1']
    worksheet2 = workbook2['Sheet1']
    worksheet3 = workbook3['Sheet1']

    worksheet4 = workbook4['Sheet1']
    worksheet5 = workbook5['Sheet1']
    worksheet6 = workbook6['Sheet1']

    # select the active worksheet as the destination worksheet
    destination_workbook = openpyxl.load_workbook('C:\\Users\\Cprickett\\YOUR-FILE-PATH')
    destination_worksheet = destination_workbook['Sheet1']

    # MOES - copy the values from the source worksheets to the destination worksheet
    if worksheet1['F26'].value is not None:
        destination_worksheet['F11'] = worksheet1['F26'].value
    else:
        print('Error: Actual Hours is empty or None.')

    if worksheet1['J26'].value is not None:
        destination_worksheet['F10'] = worksheet1['J26'].value
    else:
        print('Error: Scheduled Hours is empty or None.')

    if worksheet1['X26'].value is not None:
        destination_worksheet['F8'] = worksheet1['X26'].value
    else:
        print('Error: Labor Dollars is empty or None.')

    if worksheet2['J50'].value is not None:
        destination_worksheet['F2'] = worksheet2['J50'].value
    else:
        print('Error: Blackboard Revenues is empty or None.')
        
    if worksheet2['K50'].value is not None:
        destination_worksheet['F3'] = worksheet2['K50'].value
    else:
        print('Error: Blackboard Customer Counts is empty or None.')

    if worksheet3['F23'].value is not None:
        destination_worksheet['F14'] = worksheet3['F23'].value
    else:
        print('Error: unch Percentage is empty or None.')
    
    #EINSTEINS - copy the values from the source worksheets to the destination worksheet 
    if worksheet4['F28'].value is not None:
        destination_worksheet['F28'] = worksheet4['F28'].value
    else:
        print('Error: Actual Hours is empty or None.')

    if worksheet4['J26'].value is not None:
        destination_worksheet['F27'] = worksheet4['J26'].value
    else:
        print('Error: Scheduled Hours is empty or None.')

    if worksheet4['X28'].value is not None:
        destination_worksheet['F25'] = worksheet4['X28'].value
    else:
        print('Error: Actual Wages is empty or None.')

    if worksheet5['J34'].value is not None:
        destination_worksheet['F19'] = worksheet5['J34'].value
    else:
        print('Error: Blackboard Revenues is empty or None.')
    
    if worksheet5['K34'].value is not None:
        destination_worksheet['F20'] = worksheet5['K34'].value
    else:
        print('Error: Blackboard Customer Counts is empty or None.')

    if worksheet6['F17'].value is not None:
        destination_worksheet['F31'] = worksheet6['F17'].value
    else:
        print('Error: Punch Percentage is empty or None.')
    
    print("Data has been copied successfully.")
    
    # save the destination workbook
    destination_workbook.save('C:\\Users\\Cprickett\\YOUR-FILE-PATH')

copy_data()

print('Success!')
